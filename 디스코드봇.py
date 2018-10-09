import discord
import asyncio
import random
import openpyxl
from discord import Member
from discord.ext import commands
import youtube_dl
from urllib.request import urlopen, Request
import urllib
import bs4
import urllib.request
import bs4


client = discord.Client()
players = {}
@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("------------------")
    await client.change_presence(game=discord.Game(name='이런 화창한날에 너같은 녀석은...', type=1))



@client.event
async def on_message(message):

    if message.content.startswith('!안녕'):
        await client.send_message(message.channel, "안녕하세요")

    if message.content.startswith('!오늘배그'):
        randomNum = random.randrange(1, 3)
        if randomNum==1:
            await client.send_message(message.channel, embed=discord.Embed(title="배그각입니다.", color=discord.Color.blue()))
        else:
            await client.send_message(message.channel, embed=discord.Embed(title="자러갑시다....", color=discord.Color.red()))

    if message.content.startswith("!홋치"):
        file = openpyxl.load_workbook('기억.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A"+str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await client.send_message(message.channel, "좀더 때려줘어어엉항항하ㅑㅎ아아")
                await client.send_message(message.channel, "★ 현재 사용중인 데이터 저장용량 : 200/" + str(i)+" ★")
                break
        file.save("기억.xlsx")

    if message.content.startswith("!말해"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith("!기억 초기화") or message.content.startswith("!기억초기화"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        for i in range(1, 251):
            sheet["A"+str(i)].value = "-"
        await client.send_message(message.channel, "기억초기화 완료")
        file.save("기억.xlsx")

    if message.content.startswith("!데이터목록") or message.content.startswith("!데이터 목록"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        for i in range(1, 201):
            if sheet["A" + str(i)].value == "-" and i == 1:
                await client.send_message(message.channel, "데이터 없음")
            if sheet["A" + str(i)].value == "-":
                break
            await client.send_message(message.channel, "A : "+sheet["A" + str(i)].value + " B : "+ sheet["B" + str(i)].value)

    if message.content.startswith("!명령어"):
        channel = message.channel
        embed = discord.Embed(
            title = '명령어들이다 크크크큭',
            description = '각각의 명령어들 이다 잘 봐둬라 큭...',
            colour = discord.Colour.blue()
        )

        embed.set_footer(text = '끗')
        embed.add_field(name = '!안녕', value = '인사하고싶을때 ㄱㄱ',inline = False)
        embed.add_field(name='!오늘배그', value='오늘 배그각 알려줌', inline=False)
        embed.add_field(name='!홋치', value='!홋치 단어1 단어2 형식으로 적으면 학습함', inline=False)
        embed.add_field(name='!말해', value='!말해 단어1 형식으로 적으면 학습한내용 말함', inline=False)
        embed.add_field(name='!기억초기화', value='학습한 데이터 초기화함', inline=False)
        embed.add_field(name='!데이터목록', value='학습한 데이터목록 알려줌', inline=False)
        embed.add_field(name='!모두모여', value='모두를 언급함', inline=False)
        embed.add_field(name='!들어와', value='봇이 음성채널에 들어옴', inline=False)
        embed.add_field(name='!나가', value='봇이 음성채널에 나감', inline=False)
        embed.add_field(name='!재생', value='!재생 유튜브링크 형식으로 적으면 유튜브 틀어줌', inline=False)
        embed.add_field(name='!일시정지', value='재생중인 유튜브 일시정지함', inline=False)
        embed.add_field(name='!다시재생', value='정지중인 유튜브 다시 재생함', inline=False)
        embed.add_field(name='!멈춰', value='재생,정지중인 유튜브 없어짐(영상목록 초기화)', inline=False)
        embed.add_field(name='!날씨', value='!날씨 원하는지역 을 입력하면 그 지역의 날씨정보를 제공합니다.', inline=False)
        embed.add_field(name='!롤', value='!롤 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        embed.add_field(name='!배그솔로', value='!배그솔로 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        embed.add_field(name='!배그듀오', value='!배그듀오 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        embed.add_field(name='!배그스쿼드', value='!배그스쿼드 닉네임 형식으로 적으면 그 닉네임에대한 정보를 알려줍니다..', inline=False)
        embed.add_field(name='!고양이', value='!고양이 라고 적으면 고양이 사진이 나옵니다..', inline=False)
        embed.add_field(name='!강아지', value='!강아지 라고 적으면 강아지 사진이 나옵니다.', inline=False)
        embed.add_field(name='!네코', value='!네코 라고 적으면 2D 고양이 이미지가 나옵니다', inline=False)
        embed.add_field(name='!실시간검색어, !실검', value='!실시간검색어, !실검 이라고 적으면 네이버의 실시간 검색어 순위가 나타납니다.', inline=False)
        embed.add_field(name='!번역 번역할문자', value='!번역 번역할문자 이라고 적으면 번역할 문자를 번역한 링크가 나타납니다. ("띄어쓰기를 하시면 안됩니다. _,-등으로 구분해주세요.")', inline=False)

        await client.send_message(channel,embed=embed)

    if message.content.startswith("!모두모여"):
        await client.send_message(message.channel, "@everyone")

    if message.content.startswith("!들어와"):
        channel = message.author.voice.voice_channel
        server = message.server
        voice_client = client.voice_client_in(server)
        print("들어와")
        print(voice_client)
        print("들어와")
        if voice_client== None:
            await client.send_message(message.channel, '들어왔습니다') # 호오.... 나를 부르는건가? 네녀석.. 각오는 되있겠지?
            await client.join_voice_channel(channel)
        else:
            await client.send_message(message.channel, '봇이 이미 들어와있습니다.') # 응 이미 들어와있어 응쓰게싸




    if message.content.startswith("!나가"):
            server = message.server
            voice_client = client.voice_client_in(server)
            print("나가")
            print(voice_client)
            print("나가")
            if voice_client == None:
                await client.send_message(message.channel,'봇이 음성채널에 접속하지 않았습니다.') # 원래나가있었음 바보녀석 니녀석의 죄는 "어리석음" 이라는 .것.이.다.
                pass
            else:
                await client.send_message(message.channel, '나갑니다') # 나가드림
                await voice_client.disconnect()


    if message.content.startswith("!재생"):
        server = message.server
        voice_client = client.voice_client_in(server)
        msg1 = message.content.split(" ")
        url = msg1[1]
        player = await voice_client.create_ytdl_player(url)
        players[server.id] = player
        await client.send_message(message.channel, embed=discord.Embed(description="재생한다!!!!"))
        player.start()


    if message.content.startswith("!일시정지"):
        id = message.server.id
        await client.send_message(message.channel, embed=discord.Embed(description="장비를 정비합니다"))
        players[id].pause()

    if message.content.startswith("!다시재생"):
        id = message.server.id
        await client.send_message(message.channel, embed=discord.Embed(description="다시재생한다!!!!"))
        players[id].resume()

    if message.content.startswith("!멈춰"):
        id = message.server.id
        await client.send_message(message.channel, embed=discord.Embed(description="세계의 시간은 멈춰있다..."))
        players[id].stop()

    if message.content.startswith("!날씨"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location+'날씨')
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        todayBase = bsObj.find('div', {'class': 'main_info'})

        todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
        todayTemp = todayTemp1.text.strip()  # 온도
        print(todayTemp)

        todayValueBase = todayBase.find('ul', {'class': 'info_list'})
        todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
        todayValue = todayValue2.text.strip()  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        print(todayValue)

        todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
        todayFeelingTemp = todayFeelingTemp1.text.strip()  # 체감온도
        print(todayFeelingTemp)

        tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
        tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
        tomorrowTemp2 = tomorrowTemp1.find('dl')
        tomorrowTemp3 = tomorrowTemp2.find('dd')
        tomorrowTemp = tomorrowTemp3.text.strip()  # 내일 오전,오후온도
        print(tomorrowTemp)

        embed = discord.Embed(
            title=learn[1]+ ' 날씨 정보',
            description=learn[1]+ '날씨 정보입니다.',
            colour=discord.Colour.gold()
        )
        embed.add_field(name='현재온도', value=todayTemp+'˚', inline=False)  # 현재온도
        embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
        embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        embed.add_field(name='내일 오전/오후 날씨', value=todayFeelingTemp, inline=False)  # 내일날씨

        await client.send_message(message.channel,embed=embed)


    if message.content.startswith("!롤"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)

        url = "http://www.op.gg/summoner/userName=" + enc_location
        html = urllib.request.urlopen(url)

        bsObj = bs4.BeautifulSoup(html, "html.parser")
        rank1 = bsObj.find("div", {"class": "TierRankInfo"})
        rank2 = rank1.find("div", {"class": "TierRank"})
        rank3 = rank2.find("span", {"class": "tierRank"})
        rank4 = rank3.text  # 티어표시 (브론즈1,2,3,4,5 등등)
        print(rank4)
        if rank4 != 'Unranked':
          jumsu1 = rank1.find("div", {"class": "TierInfo"})
          jumsu2 = jumsu1.find("span", {"class": "LeaguePoints"})
          jumsu3 = jumsu2.text
          jumsu4 = jumsu3.strip()#점수표시 (11LP등등)
          print(jumsu4)

          winlose1 = jumsu1.find("span", {"class": "WinLose"})
          winlose2 = winlose1.find("span", {"class": "wins"})
          winlose2_1 = winlose1.find("span", {"class": "losses"})
          winlose2_2 = winlose1.find("span", {"class": "winratio"})

          winlose2txt = winlose2.text
          winlose2_1txt = winlose2_1.text
          winlose2_2txt = winlose2_2.text #승,패,승률 나타냄  200W 150L Win Ratio 55% 등등

          print(winlose2txt + " " + winlose2_1txt + " " + winlose2_2txt)

        channel = message.channel
        embed = discord.Embed(
            title='롤 정보',
            description='롤 정보입니다.',
            colour=discord.Colour.green()
        )
        if rank4=='Unranked':
            embed.add_field(name='당신의 티어', value=rank4, inline=False)
            embed.add_field(name='-당신은 언랭-', value="언랭은 더이상의 정보는 제공하지 않습니다.", inline=False)
            await client.send_message(channel, embed=embed)
        else:
         embed.add_field(name='당신의 티어', value=rank4, inline=False)
         embed.add_field(name='당신의 LP(점수)', value=jumsu4, inline=False)
         embed.add_field(name='당신의 승,패 정보', value=winlose2txt+" "+winlose2_1txt, inline=False)
         embed.add_field(name='당신의 승률', value=winlose2_2txt, inline=False)
         await client.send_message(channel, embed=embed)



    if message.content.startswith("!배그솔로"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/"+enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        solo1 = bsObj.find("div", {"class": "overview"})
        solo2 = solo1.text
        solo3 = solo2.strip()
        channel = message.channel
        embed = discord.Embed(
            title='배그솔로 정보',
            description='배그솔로 정보입니다.',
            colour=discord.Colour.green())
        if solo3 == "No record":
            print("솔로 경기가 없습니다.")
            embed.add_field(name='배그를 한판이라도 해주세요', value='솔로 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            solo4 = solo1.find("span", {"class": "value"})
            soloratting = solo4.text  # -------솔로레이팅---------
            solorank0_1 = solo1.find("div", {"class": "grade-info"})
            solorank0_2 = solorank0_1.text
            solorank = solorank0_2.strip()  # -------랭크(그마,브론즈)---------

            print("레이팅 : " + soloratting)
            print("등급 : " + solorank)
            print("")
            embed.add_field(name='레이팅', value=soloratting, inline=False)
            embed.add_field(name='등급', value=solorank, inline=False)

            soloKD1 = bsObj.find("div", {"class": "kd stats-item stats-top-graph"})
            soloKD2 = soloKD1.find("p", {"class": "value"})
            soloKD3 = soloKD2.text
            soloKD = soloKD3.strip()  # -------킬뎃(2.0---------
            soloSky1 = soloKD1.find("span", {"class": "top"})
            soloSky2 = soloSky1.text  # -------상위10.24%---------

            print("킬뎃 : " + soloKD)
            print("킬뎃상위 : " + soloSky2)
            print("")
            embed.add_field(name='킬뎃,킬뎃상위', value=soloKD+" "+soloSky2, inline=False)
            #embed.add_field(name='킬뎃상위', value=soloSky2, inline=False)

            soloWinRat1 = bsObj.find("div", {"class": "stats"})  # 박스
            soloWinRat2 = soloWinRat1.find("div", {"class": "winratio stats-item stats-top-graph"})
            soloWinRat3 = soloWinRat2.find("p", {"class": "value"})
            soloWinRat = soloWinRat3.text.strip()  # -------승률---------
            soloWinRatSky1 = soloWinRat2.find("span", {"class": "top"})
            soloWinRatSky = soloWinRatSky1.text.strip()  # -------상위?%---------

            print("승률 : " + soloWinRat)
            print("승률상위 : " + soloWinRatSky)
            print("")
            embed.add_field(name='승률,승률상위', value=soloWinRat+" "+soloWinRatSky, inline=False)
            #embed.add_field(name='승률상위', value=soloWinRatSky, inline=False)

            soloHead1 = soloWinRat1.find("div", {"class": "headshots stats-item stats-top-graph"})
            soloHead2 = soloHead1.find("p", {"class": "value"})
            soloHead = soloHead2.text.strip()  # -------헤드샷---------
            soloHeadSky1 = soloHead1.find("span", {"class": "top"})
            soloHeadSky = soloHeadSky1.text.strip()  # # -------상위?%---------

            print("헤드샷 : " + soloHead)
            print("헤드샷상위 : " + soloHeadSky)
            print("")
            embed.add_field(name='헤드샷,헤드샷상위', value=soloHead+" "+soloHeadSky, inline=False)
            #embed.add_field(name='헤드샷상위', value=soloHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("!배그듀오"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "duo modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그듀오 정보',
            description='배그듀오 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('듀오 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='듀오 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)


            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD+" "+duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)


    if message.content.startswith("!배그스쿼드"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "squad modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그스쿼드 정보',
            description='배그스쿼드 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('스쿼드 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='스쿼드 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)


            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD+" "+duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith('!고양이'):
        embed = discord.Embed(
            title='고양이는',
            description='멍멍',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase+str(randomNum)
        embed.set_image(url = urlF)
        await client.send_message(message.channel, embed=embed)

    if message.content.startswith('!강아지'):
        embed = discord.Embed(
            title='강아지는',
            description='야옹야옹',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240/dog?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase+str(randomNum)
        embed.set_image(url = urlF)
        await client.send_message(message.channel, embed=embed)

    if message.content.startswith('!네코'):
        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed2 = discord.Embed(
            colour=discord.Colour.green()
        )
        embed3 = discord.Embed(
            colour=discord.Colour.green()
        )
        randomnumber = random.randrange(100, 407)
        randomgiho = random.randrange(1,3)
        print('?번째사진 : '+str(randomnumber))
        print('기호 : '+str(randomgiho))
        strandomnumber = str(randomnumber)
        file1 = '.png'
        file2 = '.jpg'
        file3 = '.jpeg'
        giho = '_'
        if randomgiho==1:
            urlbase1 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file1
            urlbase2 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file2
            urlbase3 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file3
            embed.set_image(url=urlbase1)
            embed2.set_image(url=urlbase2)
            embed3.set_image(url=urlbase3)
            await client.send_message(message.channel, embed=embed)
            await client.send_message(message.channel, embed=embed2)
            await client.send_message(message.channel, embed=embed3)
        else:
            urlbase_1 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file1
            urlbase_2 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file2
            urlbase_3 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file3
            embed.set_image(url=urlbase_1)
            embed2.set_image(url=urlbase_2)
            embed3.set_image(url=urlbase_3)
            await client.send_message(message.channel, embed=embed)
            await client.send_message(message.channel, embed=embed2)
            await client.send_message(message.channel, embed=embed3)


        #https://cdn.nekos.life/neko/neko_001.png
        #https://cdn.nekos.life/neko/neko001.png

    if message.content.startswith('!구규범'):
        number = random.randrange(1,23)
        filename = 'gu'+str(number)+'.jpg'
        await client.send_message(message.channel, embed=discord.Embed(title="그의 찬란한 모습.....", color=discord.Color.red()))
        await client.send_file(message.channel, filename)


    if message.content.startswith('!실시간검색어') or message.content.startswith('!실검'):
        url = "https://www.naver.com/"
        html = urllib.request.urlopen(url)

        bsObj = bs4.BeautifulSoup(html, "html.parser")
        realTimeSerach1 = bsObj.find('div', {'class': 'ah_roll_area PM_CL_realtimeKeyword_rolling'})
        realTimeSerach2 = realTimeSerach1.find('ul', {'class': 'ah_l'})
        realTimeSerach3 = realTimeSerach2.text
        realTimeSerach4 = realTimeSerach3.strip()
        print(realTimeSerach4)
        embed = discord.Embed(
            title='네이버 실시간 검색어',
            description='실시간검색어',
            colour=discord.Colour.green()
        )
        embed.add_field(name='현재 실시간 검색어 순위입니다.', value=realTimeSerach4, inline=False)
        await client.send_message(message.channel,embed=embed)

    if message.content.startswith('!번역'):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = 'https://translate.google.co.kr/#ko/en/' + enc_location
        print(url)
        await client.send_message(message.channel, url)
        
    if message.content.startswith('!움짤'):
        embed = discord.Embed(
            title='랜덤움짤',
            description='쨜쨜쨜쨜ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랼',
            colour=discord.Colour.green()
        )
        url = "http://www.gifbin.com/random"
        urlBase = "http://www.gifbin.com"
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        gif1 = bsObj.find('form', {'id': 'share-form'})
        gif2 = gif1.find('a')
        gif3 = gif2["href"]
        gifURL = urlBase + gif3
        print(gifURL)
        embed.set_image(url=gifURL)
        await client.send_message(message.channel, embed=embed)
        #"http://www.gifbin.com/random"

    if message.content.startswith("!테스트"):
        embed = discord.Embed(
            title='테스트',
            description='ㅇㅇ',
            colour=discord.Colour.green()
        )
        embed.add_field(name='테스트중..', value="https://youtu.be/WptXk39wiIQ", inline=False)
        embed.add_field(name='테스트중..', value="https://youtu.be/2kZVEUGLgy4", inline=False)
        embed.add_field(name='테스트중..', value="https://youtu.be/z5A0joDLq38", inline=False)
        await client.send_message(message.channel, embed=embed)

        






client.run('NDk4MDYwMzA2OTQ2MTk1NDU2.DpoOrQ.DSlcQBwthAugLaHJyNY8HN_POWQ')
